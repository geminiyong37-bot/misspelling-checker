import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

FILE_COLORS = [
    "DBEAFE", "FEF3C7", "D1FAE5", "FCE7F3", "E0E7FF",
    "FED7AA", "E5E7EB", "DDD6FE", "CFFAFE", "FECACA"
]

ORANGE_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
HEADER_FILL = PatternFill(start_color="374151", end_color="374151", fill_type="solid")

def build_rows_from_doc(doc):
    sentences = doc.get("sentences", [])
    source = doc.get("file") or doc.get("source") or "알 수 없는 문서"
    
    rows = []
    for idx, sentence in enumerate(sentences):
        rows.append({
            "file": source,
            "page": sentence.get("pageNumber", 1),
            "original": sentence.get("text", ""),
            "corrected": "",
            "help": sentence.get("meta") or "본문",
            "index": idx + 1
        })
    return rows

def create_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "오타검토결과"

    headers = ["문서명", "수정전", "수정후", "비고"]
    ws.append(headers)

    # 너비 설정
    widths = [30, 40, 40, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = HEADER_FILL
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"

    # 중복 제거 및 데이터 그룹화
    unique_errors = {}
    for row in rows:
        key = (row["original"], row["corrected"])
        if key not in unique_errors:
            unique_errors[key] = {
                "file": row["file"],
                "original": row["original"],
                "corrected": row["corrected"],
                "help": row.get("help") or row.get("reason", ""),
                "count": 0,
                "pages": set()
            }
        unique_errors[key]["count"] += 1
        unique_errors[key]["pages"].add(str(row.get("page", 1)))

    def is_highlighted(original, corrected, note):
        norm_orig = (original or "").strip().replace(" ", "")
        norm_corr = (corrected or "").strip().replace(" ", "")
        
        # 순화어, 단위 관련 단순 규정 준수는 강조에서 제외
        note_str = str(note or "")
        if "순화" in note_str or "금액" in note_str or "단위" in note_str:
            return False
            
        return bool(norm_corr and norm_orig != norm_corr)

    grouped_data = list(unique_errors.values())
    grouped_data.sort(key=lambda x: (x["file"], not is_highlighted(x["original"], x["corrected"], x["help"])))

    file_color_map = {}
    color_idx = 0
    prev_file = None

    for data in grouped_data:
        original = data["original"]
        corrected = data["corrected"]
        note = data["help"]
        count = data["count"]
        display_name = os.path.basename(data["file"])
        if display_name not in file_color_map:
            file_color_map[display_name] = FILE_COLORS[color_idx % len(FILE_COLORS)]
            color_idx += 1
        
        file_bg_color = file_color_map[display_name]
        
        note_text = f"[중복 {count}회] {note}" if count > 1 else note

        ws.append([display_name, original, corrected, note_text])
        current_row = ws.max_row
        
        if prev_file is not None and prev_file != display_name:
            # 문서 교체 시 상단 테두리 추가
            thick_border = Border(top=Side(style='medium', color='000000'))
            for col in range(1, 5):
                # 기존 셀 테두리가 있을 수 있으니 Top 테두리만 새로 지정하거나 합침
                # 여기서는 단순하게 Top border 지정
                ws.cell(row=current_row, column=col).border = thick_border
        
        prev_file = display_name
        
        if is_highlighted(original, corrected, note):
            file_fill = PatternFill(start_color=file_bg_color, end_color=file_bg_color, fill_type="solid")
            ws.cell(row=current_row, column=2).fill = file_fill
            ws.cell(row=current_row, column=3).fill = file_fill

        # 안전하게 제일 마지막에 텍스트 줄바꿈 및 가운데 정렬 다시 덮어씌우기
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=current_row, column=2).alignment = center_align
        ws.cell(row=current_row, column=3).alignment = center_align

    return wb

def export_to_excel(rows, output_path):
    wb = create_workbook(rows)
    wb.save(output_path)
